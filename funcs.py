
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError, InvalidRequestError
from .models import VCode, Consigment, Collection, base_path, Table_row
from .solo_settings import replace_dict
from openpyxl import load_workbook
import random, string, time
import os


def find_code(vcode, session):
    code_in_base = session.query(VCode).filter(VCode.code.contains(vcode)).all()

    if not code_in_base:
        return False
    len_vcode = len(vcode)
    for i, code in enumerate(code_in_base.copy()):
        if len(code.code) > len_vcode+2:
            code_in_base.remove(code)
    # if len(code_in_base) > 1:
    #     return min([code.code for code in code_in_base], key=len)
    return code_in_base

def find_analog_photowp(vcode, session):
    digit_code = ""
    for lett in vcode:
        if lett.isdigit():
            digit_code += lett
        else:
            break
    response = find_code(digit_code, session)
    if not response:
        return False
    codes = []
    for _code in response.copy():
        if "V" in _code.code or "P" in _code.code:
            if len(_code.consigments) > 0 and _code.consigments[0].amount > 0:
                codes.append(_code)
    return codes

def check_for_photo(session, vcode, row):
    if "V" in vcode or "P" in vcode:
        check_photo = find_analog_photowp(vcode, session)
        if check_photo and len(check_photo) > 1:
            row.comment += "Аналоги: "
            for code in check_photo:
                if code.code == vcode:
                    continue
                row.comment += code.code + " "
                row.comment += str(code.consigments[0].amount) + " шт, "
            row.comment = row.comment[:-2]
            if row.number == 0:
                row.number = 1
        return True
    return False

def get_for_table(data_str, session):
    # data_list = data_str.split()
    data_list = data_str
    """выше не менять"""

    for_range = range(len(data_list))

    table_rows = []
    skip = False
    in_boxes = False
    boxes_num = 0
    for i in for_range:
        if skip:
            skip = False
            continue

        vcode = data_list[i]
        if i == for_range.stop-1 or len(data_list[i+1]) > 3:
            in_boxes = True
            # skip = True
        elif "*" in data_list[i+1]:
            in_boxes = True
            skip = True
            # boxes_num = int(data_list[i+1][1:])
            boxes_num = data_list[i+1].replace("*", "")
        else:
            in_boxes = False
            skip = True

        if not in_boxes:
            number = data_list[i + 1]
        else:
            number = 0

        # if len(vcode) < 4 and len(str(number)) > 3:
        #     vcode, number = number, vcode

        number = int(number)
        vcode = vcode.upper()
        for lett in vcode:
            if lett in replace_dict:
                vcode = vcode.replace(lett, replace_dict[lett])
        row = Table_row(vcode, number)
        table_rows.append(row)

    # for row in table_rows:
        # print('trying %s\n' % row.vcode)
        code_in_base = session.query(VCode).filter(VCode.code == row.vcode).first()
        _code_in_base = None
        if not code_in_base:
            row.comment += "арт. %s не найден " % row.vcode
            check_for_photo(session, vcode, row)
                    # continue
            code_in_base = find_code(vcode, session)
            if not code_in_base:
                row.consig = ""
                # row.comment += "Артикул не найден "
                continue

            elif len(code_in_base) > 1:
                minimal = min([code.code for code in code_in_base], key=len) # самый короткий артикул
                row.comment += "Возможные артикулы: "
                for _code in code_in_base:
                    row.comment += "{} ".format(_code.code)
                    if _code.code == minimal:
                        _code_in_base = _code   # выбор самого короткого артикула
                    # row.vcode = code_in_base[0].code
                    continue
            else:
                # row.comment += "Возможно нужен %s " % code_in_base[0].code
                row.vcode = code_in_base[0].code
                code_in_base = code_in_base[0]
            # continue

        if _code_in_base:
            code_in_base = _code_in_base
            row.vcode = code_in_base.code

        if in_boxes:
            if code_in_base.collection:
                if boxes_num:
                    row.number = code_in_base.collection.boxes * int(boxes_num)
                    boxes_num = 0
                else:
                    row.number = code_in_base.collection.boxes
            else:
                row.number = 1
        if code_in_base.consigments:
            enough = False
            for consig in code_in_base.consigments:
                if consig.amount >= row.number:
                    enough = True
                    if consig.name:
                        row.consig = consig.name
                    break
            if not enough:
                row.comment += "Недостаточно на св. остатке."
                add_comment = " в наличии есть партии: "
                max_num_consig = None
                max_num = 0
                for consig in code_in_base.consigments:
                    if consig.amount >= 0:
                        if consig.amount > max_num:
                            max_num_consig = consig.name
                            max_num = consig.amount
                        add_comment += "{} {} рул. ".format(consig.name, consig.amount)
                row.consig = max_num_consig
                if len(add_comment) > 24:
                    row.comment += add_comment
        else:
            if not check_for_photo(session, vcode, row):
                row.consig = "Общая"
                row.comment = "Партий не найдено"
            else:
                continue

        check_for_photo(session, vcode, row)

    return table_rows


def read_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[0].value.split()[0]
        consig = row[1].value
        amount = row[2].value
        xlxs_rows.append((i, vcode, consig, amount))
    wb.close()
    return xlxs_rows

def read_abc_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i < 6:
            continue
        vcode = row[0].value
        name = row[1].value
        collection = row[2].value
        box = row[3].value
        xlxs_rows.append((vcode, name, collection, box))
    wb.close()
    return xlxs_rows

def get_all_from_base(xlxs_rows, session):

    codes = session.query(VCode).all()
    consigs = session.query(Consigment).all()
    

    changed_positions = []
    rows_to_check = len(consigs)
    needed_collections = ['need to add collections to vendor codes:\n',]
    need_commit = False
    for row in xlxs_rows:
        
        i = row[0]
        vcode = row[1]
        consig = row[2]
        amount = row[3]

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                break
        if not code_in_base:
            session.add(VCode(vcode))
            session.commit()
            need_commit = False
            code_in_base = session.query(VCode).filter(VCode.code == vcode).first()
            changed_positions.append("add new vcode: {}".format(code_in_base.code))
            codes.append(code_in_base)
            needed_collections.append(vcode)

        consig_in_base = None
        for _consig in code_in_base.consigments:
            if _consig.name == consig:
                consig_in_base = _consig
                consigs.remove(_consig)
                break     
        if not consig_in_base:
            _consig = Consigment(name=consig, vcode=code_in_base.id, amount=amount)
            session.add(_consig)
            # session.commit()
            need_commit = True
            consig_in_base = session.query(Consigment).filter(
                                                Consigment.vcode_id == code_in_base.id).filter(
                                                    Consigment.name == consig).first()
            changed_positions.append("add new consigment for {}: {}".format(code_in_base.code, 
                                                                            consig_in_base.name))
        
        if amount != consig_in_base.amount:
            changed_positions.append("vcode {} consig {} amount change from {} to {}".format(
                                                                                    code_in_base.code,
                                                                                    consig_in_base.name,
                                                                                    consig_in_base.amount,
                                                                                    amount))
            consig_in_base.amount = amount

        if need_commit and not i % 100:
            session.commit()
            need_commit = False

    for consig_in_base in consigs:
        session.delete(consig_in_base)

    session.commit()

    checked_rows = i
    if not changed_positions:
        changed_positions.append("no changes")

    if len(needed_collections) == 1:
        needed_collections = []

    return changed_positions, rows_to_check, checked_rows, needed_collections


def add_boxes_to_vcodes(xlxs_rows, session):
    """Выгрузка из ABC"""
    codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = row[0]
        name = row[1]
        collection = row[2]
        box = row[3]

        collection_in_base = None
        for _coll in collections:
            if _coll.name == collection:
                collection_in_base = _coll
                break
        if not collection_in_base:
            response.append("added collection %s" % collection)
            session.add(Collection(collection, box))
            session.commit()
            collection_in_base = session.query(Collection).filter(Collection.name == collection).first()
            collections.append(collection_in_base)

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                if not code_in_base.collection_id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("added collection to %s" % vcode)
                break
        if not code_in_base:
            response.append("added vcode %s" % vcode)
            session.add(VCode(vcode, collection_in_base.id))
            need_commit = True

        if not i % 100 and need_commit:
            session.commit()
            need_commit = False

    if need_commit:
        session.commit()

    if not response:
        response.append("no changes")
        
    return response
