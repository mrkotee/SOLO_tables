

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError, InvalidRequestError
try:
    from .sng_models import base_path, VCode, Client, CodeName
except:
    from sng_models import base_path, VCode, Client, CodeName
from openpyxl import load_workbook


def read_xlxs(xlxs_path, table_type=3):
    """
    xlxs path
    table type - 1 = торг-12, 2 = упд, 3 = спец-ия или ??
    :return:
    """

    wb = load_workbook(filename=xlxs_path)
    ws = wb.active

    code_column_id = 0
    name_column_id = 0
    start_row = 0
    if table_type == 3:
        for i, row in enumerate(ws.iter_rows()):
            if code_column_id and name_column_id:
                start_row = i
                break
            for cell in row:
                if cell.value:
                    if 'наименование' in str(cell.value).lower():
                        name_column_id = cell.column
                    elif 'артикул' in str(cell.value).lower():
                        code_column_id = cell.column

    elif table_type == 1:  # обрезать партии из наименования (у фотиков кол-во частей вместо партии)
        pass

    elif table_type == 2:  # обрезать партии из наименования (у фотиков кол-во частей вместо партии)
        pass

    else:
        return False

    code_names_dict = {}
    for i, row in enumerate(ws.rows):
        if i < start_row:
            continue

        vcode = str(ws.cell(
            row=i+1, column=code_column_id).value).strip()
        code_name = ws.cell(
            row=i+1, column=name_column_id).value

        code_names_dict[vcode] = code_name

    return code_names_dict


def change_names_xlxs(session, xlxs_path, new_path, client):
    """
    xlxs path
    table type?
    client
    :return:
    """

    client_in_base = session.query(Client).filter(Client.name == client).first()
    code_names = session.query(CodeName).all()
    code_names = []
    for code_name in client_in_base.code_names:
        for _code_name in code_name.vendor_code.names:
            code_names.append(_code_name)
    wb = load_workbook(filename=xlxs_path)
    ws = wb.active

    names_column = 0
    for row in ws.iter_rows():
        if names_column:
            break
        for cell in row:
            if names_column:
                break
            if type(cell.value) == str:
                for code_name in code_names:
                    if code_name.name in cell.value:
                        names_column = cell.column
                        break

    if names_column:
        for row_id in range(1, ws.max_row):
            if not type(ws.cell(row_id, names_column).value) == str:
                continue
            value = ws.cell(row_id, names_column).value
            for code_name in code_names:
                if code_name.name in value:
                    for client_code_name in client_in_base.code_names:
                        if client_code_name.vendor_code == code_name.vendor_code:
                            value = value.replace(code_name.name, client_code_name.name)
                            ws.cell(row_id, names_column).value = value
                            break
                    break
    else:
        for row_id in range(1, ws.max_row):
            for col_id in range(1, ws.max_column):
                if not type(ws.cell(row_id, col_id).value) == str:
                    continue
                if len(ws.cell(row_id, col_id).value) < 50:
                    continue

                value = ws.cell(row_id, col_id).value
                for code_name in code_names:
                    if code_name.name in value:
                        for client_code_name in client_in_base.code_names:
                            if client_code_name.vendor_code == code_name.vendor_code:
                                value = value.replace(code_name.name, client_code_name.name)
                                ws.cell(row_id, col_id).value = value
                                break
                        break

    wb.save(new_path)
    wb.close()


def add_names_to_base(session, code_names_dict, client=None, change_existed=False):
    """
    dictionary {vcode: code_name}
    client (optional)
    change_existed (optional) - изменять уже добавленные клиенту наименования
    :return:
    """

    if client:
        client_in_base = session.query(Client).filter(Client.name == client).first()
        if not client_in_base:
            session.add(Client(client))
            session.commit()
            client_in_base = session.query(Client).filter(Client.name == client).first()

    for code, code_name in code_names_dict.items():
        base_code = session.query(VCode).filter(VCode.code == code).first()
        if not base_code:
            session.add(VCode(code))
            session.commit()
            base_code = session.query(VCode).filter(VCode.code == code).first()

        code_name_in_base = None
        for _code_name in base_code.names:
            if code_name == _code_name.name:
                code_name_in_base = _code_name
                break
        if code_name_in_base:
            if client and client_in_base not in code_name_in_base.clients:
                client_has_code = False
                for code_name in base_code.names:
                    if code_name in client_in_base.code_names:
                        client_has_code = True
                if not client_has_code:
                    code_name_in_base.clients.append(client_in_base)
                    session.commit()
            continue

        else:
            code_name_in_base = CodeName(code_name, base_code.id)
            if client:
                code_name_in_base.clients.append(client_in_base)
            session.add(code_name_in_base)

            session.commit()
