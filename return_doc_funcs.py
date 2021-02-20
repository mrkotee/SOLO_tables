import os
from datetime import datetime as dt
from openpyxl import load_workbook
import decimal
try:
    from .num2str import num2text, decimal2text
    from .solo_settings import mounth_names, solo_path
except:
    from num2str import num2text, decimal2text
    from solo_settings import mounth_names, solo_path


def read_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb.active

    # for row in ws.iter_rows():
    #     for cell in row:
    #         if cell.value:
    #             print("r{} c{} {}".format(cell.row, cell.column, cell.value))
    positions_list = get_all_positions(ws)
    wb.close()
    return positions_list



def get_doc_of_return(filepath, returned_vcodes_dict, filepath_to_save=None, nds=False):
    wb = load_workbook(filename=filepath)
    ws = wb.active

    change_cap(ws)
    write_new_table(ws, returned_vcodes_dict, nds)

    if filepath_to_save:
        wb.save(filepath_to_save)
    else:
        wb.save(filepath)
    wb.close()



def change_cap(work_sheet, new_doc_id=0, new_doc_date=None):
    """
    r2 c1 vООО Альянс Декор, ИН
    r* c39 - коды  (удалить)
    r7 c3 vООО КОМПАНИЯ "СПАРТАК-ОМ", ИНН
    r9 c3 v610001, Кировская обл, Киров г,   (удалить)
    r11 c3 vООО Альянс Декор, ИНН
    r13 c3 vООО КОМПАНИЯ "СПАРТАК-ОМ", ИНН
    r18 c12 v201123/33  - номер документа
    r18 c15 v23.11.2020  - дата документа
    """
    first_firm_name = work_sheet.cell(row=3, column=2).value
    second_firm_name = work_sheet.cell(row=8, column=4).value

    work_sheet.cell(row=10, column=4).value = ''  # удаление адреса доставки
    # замена первой фирмы на вторую
    work_sheet.cell(row=3, column=2).value, work_sheet.cell(row=12, column=4).value = second_firm_name, second_firm_name
    # замена второй фирмы на первую
    work_sheet.cell(row=8, column=4).value, work_sheet.cell(row=14, column=4).value = first_firm_name, first_firm_name

    if new_doc_id:
        work_sheet.cell(row=19, column=13).value = new_doc_id
    else:
        work_sheet.cell(row=19, column=13).value = ''

    if new_doc_date:
        work_sheet.cell(row=19, column=16).value = new_doc_date  # new date  %dd.%mm.%YYYY
    else:
        work_sheet.cell(row=19, column=16).value = "{:02d}.{:02d}.{}".format(dt.now().day, dt.now().month,
                                                                             dt.now().year)

    for i, row in enumerate(work_sheet.iter_rows()):  # удаление кодов
        if i < 2:
            continue
        if row[39].value:
            row[39].value = ''


def write_new_table(work_sheet, returned_vcodes_dict, nds):
    def find_cells_cord(text):
        # return cords = (Row, Column)
        finded_cords = []
        for row in work_sheet.iter_rows():
            for cell in row:
                if type(cell.value) == str:
                    if text in cell.value:
                        finded_cords.append((cell.row, cell.column))
        return finded_cords

    def change_positon_data(row_id, returned_vcodes_dict, nds, returned_rows_id):
        position = {}
        position['vcode_full'] = work_sheet.cell(row=row_id, column=3).value
        position['vcode'] = work_sheet.cell(row=row_id, column=7).value.upper()
        position['number'] = int(work_sheet.cell(row=row_id, column=17).value)
        position['weight'] = float(work_sheet.cell(row=row_id, column=21).value)
        position['price'] = float(work_sheet.cell(row=row_id, column=26).value)
        position['summ'] = float(work_sheet.cell(row=row_id, column=39).value)
        if nds:
            position['summ_nds'] = float(work_sheet.cell(row=row_id, column=36).value)

        for returned_vcode, returned_vcode_num in returned_vcodes_dict.items():
            returned_vcode_num = int(returned_vcode_num)
            if str(position['vcode']) == str(returned_vcode):
                returned_rows_id += 1
                work_sheet.cell(row=row_id, column=2).value = returned_rows_id

                if not nds:
                    position['price'] = position['summ']/position['number']
                    work_sheet.cell(row=row_id, column=26).value = position['price']
                    work_sheet.cell(row=row_id, column=29).value = position['summ']
                    work_sheet.cell(row=row_id, column=35).value = ''
                    work_sheet.cell(row=row_id, column=36).value = ''

                if int(position['number']) != int(returned_vcode_num):
                    new_weight = (float(position['weight']) / int(position['number'])) * int(
                        returned_vcode_num)
                    work_sheet.cell(row=row_id, column=21).value = new_weight
                    position['weight'] = new_weight
                    new_summ = (float(position['summ']) / int(position['number'])) * int(
                        returned_vcode_num)
                    work_sheet.cell(row=row_id, column=39).value = new_summ
                    position['summ'] = new_summ
                    if not nds:
                        work_sheet.cell(row=row_id, column=29).value = new_summ
                    else:
                        new_summ_nds = (float(position['summ_nds']) / int(
                            position['number'])) * int(
                            returned_vcode_num)
                        work_sheet.cell(row=row_id, column=36).value = new_summ_nds
                        position['summ_nds'] = new_summ_nds

                    work_sheet.cell(row=row_id, column=17).value = int(returned_vcode_num)
                    work_sheet.cell(row=row_id, column=24).value = int(returned_vcode_num)
                    position['number'] = returned_vcode_num
    
                return position

        for col_id in range(work_sheet.max_column):
            if type(work_sheet.cell(row=row_id, column=col_id+1)).__name__ == 'MergedCell':
                continue
            work_sheet.cell(row=row_id, column=col_id+1).value = ''
        return None

    def change_sum(row_id, sum_number, sum_weight, sum_summ, sum_without_nds, sum_nds=''):
        work_sheet.cell(row=row_id, column=17).value = sum_number
        work_sheet.cell(row=row_id, column=24).value = sum_number
        work_sheet.cell(row=row_id, column=21).value = sum_weight
        work_sheet.cell(row=row_id, column=36).value = sum_nds
        work_sheet.cell(row=row_id, column=39).value = sum_summ
        work_sheet.cell(row=row_id, column=29).value = sum_without_nds

    table_start_cord = find_cells_cord('Но-')
    table_end_cord = find_cells_cord('Итого')

    positions_list = []
    for i, table_cords in enumerate(table_start_cord):
        for row_id in range(table_cords[0] + 3, table_end_cord[i][0]):
            position = change_positon_data(row_id, returned_vcodes_dict, nds, len(positions_list))
            if position:
                positions_list.append(position)

    sum_number = sum([pos['number'] for pos in positions_list])
    sum_weight = sum([pos['weight'] for pos in positions_list])
    sum_summ = sum([pos['summ'] for pos in positions_list])
    if nds:
        sum_nds = sum([pos['summ_nds'] for pos in positions_list])
        sum_without_nds = sum_summ - sum_nds
    else:
        sum_nds = ''
        sum_without_nds = sum_summ

    final_end_cord = find_cells_cord('Всего по накладной')[0]
    change_sum(final_end_cord[0], sum_number, sum_weight, sum_summ, sum_without_nds, sum_nds)
    change_sum(table_end_cord[-1][0], sum_number, sum_weight, sum_summ, sum_without_nds, sum_nds)

    '''    r34 c4 vи содержит
    r34 c5 vДва  - кол-во строк (позиций)
    r39 c5 Всего мест
    r37 c6 Четыре
    r36 c10 v     Масса груза (нетто)
    r36 c18 vДесять кг семьсот тридцать один г
    r38 c10 v     Масса груза (брутто)
    r38 c18 vДесять кг семьсот сорок г
    '''
    int_units = ((u'кг', u'кг', u'кг'), 'm')
    exp_units = ((u'г', u'г', u'г'), 'm')
    footer_start_cord = find_cells_cord('и содержит')[0]
    work_sheet.cell(row=footer_start_cord[0], column=6).value = num2text(len(positions_list)).capitalize()
    work_sheet.cell(row=footer_start_cord[0]+2, column=6).value = num2text(sum_number).capitalize()
    work_sheet.cell(row=footer_start_cord[0]+2, column=19).value = ''
    work_sheet.cell(row=footer_start_cord[0]+4, column=19).value = decimal2text(decimal.Decimal(sum_weight),
                                                                                places=3,
                                                                                int_units=int_units,
                                                                                exp_units=exp_units).capitalize()


    '''
        __________
    r44 c2 Всего отпущено  на сумму
    r45 c2 Шесть тысяч двести три рубля 52 копейки
    __________
    r46 c4 vГенеральный директор
    r46 c10 vБалабко И. Н.
    r48 c1 vГлавный (старший) бухгалтер
    r48 c10 vБородинова О. Н.
    r49 c4 vКладовщик
    r50 c10 vШереметьев Д.М.
        r53 c6 v"23"
    r53 c7 vноября
    r53 c9 v2020 года'''

    footer_start_cord = find_cells_cord('Всего отпущено  на сумму')[0]

    int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
    exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
    work_sheet.cell(row=footer_start_cord[0]+1, column=2).value = decimal2text(decimal.Decimal(sum_summ),
                                                                                places=2,
                                                                                int_units=int_units,
                                                                                exp_units=exp_units).capitalize()
    work_sheet.cell(row=footer_start_cord[0] + 3, column=11).value = ''
    work_sheet.cell(row=footer_start_cord[0] + 5, column=11).value = ''
    work_sheet.cell(row=footer_start_cord[0] + 7, column=11).value = ''
    work_sheet.cell(row=footer_start_cord[0] + 3, column=5).value = ''
    work_sheet.cell(row=footer_start_cord[0] + 6, column=5).value = ''

    work_sheet.cell(row=footer_start_cord[0] + 10, column=7).value = '"{:02d}"'.format(dt.now().day)
    work_sheet.cell(row=footer_start_cord[0] + 10, column=8).value = mounth_names[dt.now().month]
    work_sheet.cell(row=footer_start_cord[0] + 10, column=10).value = '{} года'.format(dt.now().year)


def get_all_positions(work_sheet):
    """
    r21 c1 vНо-мер    - начало таблицы
    ______строка позиции_____
    r24 c1 v1   - порядковый номер
    r24 c2 v240499/Обои - номенклатура
    r24 c6 v240499   - артикул
    r24 c16 v3   - количество
    r24 c20 v8.07   - суммарный вес
    r24 c23 v3   - кол-во повтор
    r24 c25 v1292.4  - цена (делать полной если ИП с 0% ндс)
    r24 c28 v3877.2   - сумма без ндс (делать полной если ИП с 0% ндс)
    r24 c34 v20%    - НДС (убрать если ИП)
    r24 c35 v775.44  - сумма НДС (убрать если ИП)
    r24 c38 v4652.64  - сумма с учетом ндс
    """

    def find_cells_cord(text):
        # return cords = (Row, Column)
        finded_cords = []
        for row in work_sheet.iter_rows():
            for cell in row:
                if type(cell.value) == str:
                    if text in cell.value:
                        finded_cords.append((cell.row, cell.column))
        return finded_cords

    def get_positon_data(row_id):
        position = {}
        position['vcode_full'] = work_sheet.cell(row=row_id, column=3).value
        position['vcode'] = work_sheet.cell(row=row_id, column=7).value
        position['number'] = work_sheet.cell(row=row_id, column=17).value
        position['weight'] = work_sheet.cell(row=row_id, column=21).value
        position['price'] = work_sheet.cell(row=row_id, column=26).value
        position['summ'] = work_sheet.cell(row=row_id, column=39).value
        return position

    table_start_cord = find_cells_cord('Но-')
    table_end_cord = find_cells_cord('Итого')

    positions_list = []
    for i, table_cords in enumerate(table_start_cord):
        for row_id in range(table_cords[0] + 3, table_end_cord[i][0]):
            positions_list.append(get_positon_data(row_id))

    return positions_list

