
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError, InvalidRequestError
try:
    from alchemy_models import VCode, Consigment, Collection, base_path, Table_row, \
    VCodeName, Collection_Factory, Factory
    from solo_settings import replace_dict, name_base_path, sep_files_dir, exclusive_collections, xlxs_filepath
except:
    from solo.alchemy_models import VCode, Consigment, Collection, base_path, Table_row, \
    VCodeName, Collection_Factory, Factory
    from solo.solo_settings import replace_dict, name_base_path, sep_files_dir, exclusive_collections, xlxs_filepath

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import random, string, time
import os, base64
from exchangelib import Credentials, Account, Configuration, DELEGATE, Message, Mailbox, \
  FileAttachment, HTMLBody
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate



def find_code(vcode, session):
    code_in_base = session.query(VCode).filter(VCode.code.contains(vcode)).all()

    if not code_in_base:
        return False
    len_vcode = len(vcode)
    for i, code in enumerate(code_in_base.copy()):
        if code.code[0].isdigit():  # len(code.code) > len_vcode+2 and 
            code_in_base.remove(code)
    # if len(code_in_base) > 1:
    #     return min([code.code for code in code_in_base], key=len)
    return code_in_base

def find_analog_photowp(vcode, session):
    digit_code = ""
    for lett in vcode:
        if lett.isdigit():
            digit_code += lett
        else:
            break
    response = find_code(digit_code, session)
    if not response:
        return False
    codes = []
    for _code in response.copy():
        if "V" in _code.code or "P" in _code.code:
            if len(_code.consigments) > 0 and _code.consigments[0].amount > 0:
                codes.append(_code)
    return codes

def check_for_photo(session, vcode, row):
    if "V" in vcode or "P" in vcode:
        check_photo = find_analog_photowp(vcode, session)
        if check_photo and len(check_photo) > 1:
            row.comment += " Аналоги: "
            for code in check_photo:
                if code.code == vcode:
                    continue
                row.comment += code.code + " "
                row.comment += str(code.consigments[0].amount) + " шт, "
            row.comment = row.comment[:-2]
            if row.number == 0:
                row.number = 1
        return True
    return False

def get_for_table(data_str, session, all_boxes_num=0, uni_boxes_num=0,
                    get_all_consigs=False):
    # data_list = data_str.split()
    data_list = data_str
    """выше не менять"""

    for_range = range(len(data_list))

    table_rows = []
    skip = False
    in_boxes = False
    boxes_num = 0
    for i in for_range:
        if skip:
            skip = False
            continue

        vcode = data_list[i]

        if len(vcode) < 4 and not vcode[0].isdigit():
            data_list[i+1] = vcode + data_list[i+1]
            continue

        if i == for_range.stop-1 or len(data_list[i+1]) > 3:
            in_boxes = True
            # skip = True
        elif "*" in data_list[i+1]:
            in_boxes = True
            skip = True
            # boxes_num = int(data_list[i+1][1:])
            boxes_num = data_list[i+1].replace("*", "")
        elif len(data_list[i+1]) < 4 and not data_list[i+1][0].isdigit():
            skip = True
            data_list[i+2] = data_list[i+1] + data_list[i+2]
        else:
            in_boxes = False
            skip = True

        if not in_boxes:
            number = data_list[i + 1]
        else:
            number = 0

        # if len(vcode) < 4 and len(str(number)) > 3:
        #     vcode, number = number, vcode

        number = int(number)
        vcode = vcode.upper()
        for lett in vcode:
            if lett in replace_dict:
                vcode = vcode.replace(lett, replace_dict[lett])
        row = Table_row(vcode, number)
        table_rows.append(row)

    # for row in table_rows:
        # print('trying %s\n' % row.vcode)
        code_in_base = session.query(VCode).filter(VCode.code == row.vcode).first()
        _code_in_base = None
        if not code_in_base:
            check_for_photo(session, vcode, row)
                    # continue
            code_in_base = find_code(vcode, session)
            if not code_in_base:
                row.consig = ""
                row.comment += "Артикул не найден "
                continue

            elif len(code_in_base) > 1:
                row.comment += "арт. %s не найден " % row.vcode
                minimal = min([code.code for code in code_in_base], key=len) # самый короткий артикул
                row.comment += "Возможные артикулы: "
                for _code in code_in_base:
                    row.comment += "{} ".format(_code.code)
                    if _code.code == minimal:
                        _code_in_base = _code   # выбор самого короткого артикула
                    # row.vcode = code_in_base[0].code
                    continue
            else:
                # row.comment += "Возможно нужен %s " % code_in_base[0].code
                row.vcode = code_in_base[0].code
                code_in_base = code_in_base[0]
            # continue

        if _code_in_base:
            code_in_base = _code_in_base
            row.vcode = code_in_base.code

        if in_boxes:
            if code_in_base.collection and not check_for_photo(session, vcode, row):
                if boxes_num:
                    row.number = code_in_base.collection.boxes * int(boxes_num)
                    boxes_num = 0
                elif uni_boxes_num and not code_in_base.motive:
                    row.number = code_in_base.collection.boxes * int(uni_boxes_num)
                elif all_boxes_num:
                    row.number = code_in_base.collection.boxes * int(all_boxes_num)
                else:
                    row.number = code_in_base.collection.boxes
            else:
                row.number = 1
        if code_in_base.consigments:
            enough = False
            for consig in code_in_base.consigments:
                if consig.amount >= row.number:
                    enough = True
                    if consig.name:
                        row.consig = consig.name
                    break
            if not enough:
                if get_all_consigs:
                    _num = int(row.number)
                    _consigs = sorted([cons for cons in code_in_base.consigments], 
                                        reverse=True, key=lambda cc: cc.amount)
                    for _consig in _consigs:
                        if _consig.amount > row.number:
                            row.consig = _consig.name
                            break
                        else:
                            if in_boxes:
                                rolls_in_box = code_in_base.collection.boxes
                                _num_from_consig = (_consig.amount // rolls_in_box) * rolls_in_box
                                if _num_from_consig == 0:
                                    continue
                                _num -= _num_from_consig
                                row.number = _num_from_consig
                            else:
                                _num -= _consig.amount
                                row.number = _consig.amount
                            row.comment += "Недостаточно на св. остатке."
                            row.consig = _consig.name

                            row = Table_row(row.vcode, _num, consig='Общая')
                            table_rows.append(row)

                    row.comment += "Недостаточно на св. остатке."

                else:
                    row.comment += "Недостаточно на св. остатке."
                    add_comment = " в наличии есть партии: "
                    max_num_consig = None
                    max_num = 0
                    for consig in code_in_base.consigments:
                        if consig.amount >= 0:
                            if consig.amount > max_num:
                                max_num_consig = consig.name
                                max_num = consig.amount
                            add_comment += "{} {} рул. ".format(consig.name, consig.amount)
                    row.consig = max_num_consig
                    if len(add_comment) > 24:
                        row.comment += add_comment
               
                if not row.consig:
                    row.consig = 'Общая'
        else:
            row.comment = "Партий не найдено"
            if not check_for_photo(session, vcode, row):
                row.consig = "Общая"
            else:
                continue

        check_for_photo(session, vcode, row)

    return table_rows


def read_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[0].value.split()[0]
        consig = row[1].value
        amount = row[2].value
        if not amount:
            continue
        xlxs_rows.append((i, vcode, consig, amount))
    wb.close()
    return xlxs_rows

def read_abc_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i < 6:
            continue
        vcode = row[0].value
        name = row[1].value
        collection = row[2].value
        box = row[3].value
        factory = row[8].value
        xlxs_rows.append((vcode, name, collection, box, factory))
    wb.close()
    return xlxs_rows

def get_all_from_base(xlxs_rows, session):

    codes = session.query(VCode).all()
    consigs = session.query(Consigment).all()
    

    changed_positions = []
    rows_to_check = len(consigs)
    needed_collections = ['need to add collections to vendor codes:\n',]
    need_commit = False
    for row in xlxs_rows:
        
        i = row[0]
        vcode = row[1]
        consig = row[2]
        amount = row[3]

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                break
        if not code_in_base:
            try:
                session.add(VCode(vcode))
                session.commit()
                need_commit = False
            except Exception as e:
                needed_collections.insert(0, 'ERROR: {}'.format(e))
                
            code_in_base = session.query(VCode).filter(VCode.code == vcode).first()
            changed_positions.append("add new vcode: {}".format(code_in_base.code))
            codes.append(code_in_base)
            needed_collections.append(vcode)

        consig_in_base = None
        for _consig in code_in_base.consigments:
            if _consig.name == consig:
                consig_in_base = _consig
                try:
                    consigs.remove(_consig)
                except Exception as e:
                    print(_consig, vcode, e)
                    # raise Exception
                break     
        if not consig_in_base:
            _consig = Consigment(name=consig, vcode=code_in_base.id, amount=amount)
            session.add(_consig)
            # session.commit()
            need_commit = True
            consig_in_base = session.query(Consigment).filter(
                                                Consigment.vcode_id == code_in_base.id).filter(
                                                    Consigment.name == consig).first()
            changed_positions.append("add new consigment for {}: {}".format(code_in_base.code, 
                                                                            consig_in_base.name))
        
        if amount != consig_in_base.amount:
            changed_positions.append("vcode {} consig {} amount change from {} to {}".format(
                                                                                    code_in_base.code,
                                                                                    consig_in_base.name,
                                                                                    consig_in_base.amount,
                                                                                    amount))
            consig_in_base.amount = amount

        if need_commit and not i % 100:
            session.commit()
            need_commit = False

    for consig_in_base in consigs:
        session.delete(consig_in_base)

    session.commit()

    checked_rows = i
    if not changed_positions:
        changed_positions.append("no changes")

    if len(needed_collections) == 1:
        needed_collections = []

    return changed_positions, rows_to_check, checked_rows, needed_collections


def add_boxes_to_vcodes(xlxs_rows, session):
    """Выгрузка из ABC"""

    def find_collection(collection_name, collections):
        for _coll in collections:
            if _coll.name == collection_name:
                return _coll
        return None
        
    codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = str(row[0])
        name = row[1]
        collection = row[2]
        box = row[3]

        collection_in_base = find_collection(collection, collections)

        if collection_in_base and collection_in_base.id != 125:
            if box and box != collection_in_base.boxes:
                collection = collection + "_" + str(box)
                collection_in_base = find_collection(collection, collections)

        if not collection_in_base:
            response.append("added collection %s" % collection)
            session.add(Collection(collection, box))
            session.commit()
            collection_in_base = session.query(Collection).filter(Collection.name == collection).first()
            collections.append(collection_in_base)

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                if not code_in_base.collection_id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("added collection to %s" % vcode)
                elif int(code_in_base.collection_id) != collection_in_base.id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("changed collection to %s" % vcode)
                break
        if not code_in_base:
            response.append("added vcode %s" % vcode)
            session.add(VCode(vcode, collection_in_base.id))
            need_commit = True

        if not i % 100 and need_commit:
            session.commit()
            need_commit = False

    if need_commit:
        session.commit()

    if not response:
        response.append("no changes")
        
    return response


def add_names_to_vcodes(xlxs_rows, session):
    """Выгрузка из ABC""" # to names base

    def find_element(name, _list):
        for _coll in _list:
            if _coll.name == name:
                return _coll
        return None

    codes = session.query(VCodeName).all()
    collections = session.query(Collection_Factory).all()
    factories = session.query(Factory).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = str(row[0])
        name = row[1]
        collection = row[2]
        box = row[3]
        factory = row[4]

        factory_in_base = find_element(factory, factories)

        if not factory_in_base:
            response.append("added factory %s" % factory)
            session.add(Factory(factory))
            session.commit()
            factory_in_base = session.query(Factory).filter(Factory.name == factory).first()
            factories.append(factory_in_base)

        collection_in_base = find_element(collection, collections)

        if collection_in_base and collection_in_base.id != 125:
            if box and box != collection_in_base.boxes:
                collection = collection + "_" + str(box)
                collection_in_base = find_element(collection, collections)

        if not collection_in_base:
            response.append("added collection %s" % collection)
            session.add(Collection_Factory(collection, box, factory_id=factory_in_base.id))
            session.commit()
            collection_in_base = session.query(Collection_Factory).filter(Collection_Factory.name == collection).first()
            collections.append(collection_in_base)

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                if not code_in_base.collection_id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("added collection to %s" % vcode)
                elif int(code_in_base.collection_id) != collection_in_base.id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("changed collection to %s" % vcode)
                if not code_in_base.name:
                    code_in_base.name = name
                break
        if not code_in_base:
            # response.append("added vcode %s" % vcode)
            session.add(VCodeName(vcode, collection_in_base.id, name=name))
            need_commit = True

        if not i % 100 and need_commit:
            session.commit()
            need_commit = False

    if need_commit:
        session.commit()

    if not response:
        response.append("no changes")
        
    return response


def separate_by_factories(dir_path, session, names_session):

    def get_vcodes_list(factory_name):
        result = []
        for collection in _factories_collection_pair[factory_name]:
            for vcode in collection.vcodes:
                try:
                    vcode_name = names_session.query(VCodeName).filter(VCodeName.code == vcode.code).first().name
                except AttributeError:
                    print(vcode.code)
                    vcode_name = ''
                if vcode.consigments:
                    for consig in vcode.consigments:
                        result.append((vcode.code, vcode_name, consig.name, consig.amount))
                else:
                    result.append((vcode.code, vcode_name, 'нет в наличии', 0))
        return result
    
    def create_xlsx_table(filename, data_list):

        title_style = NamedStyle(name="title")
        title_style.fill = PatternFill("solid", fgColor="00F4ECC5")
        title_style.font = Font(name='Arial', sz=10.0)

        default_style = NamedStyle(name="default")
        default_style.font = Font(name='Arial', sz=8.0)

        border = Side(style='thin', color='00CCC085')
        borders = Border(left=border, top=border, right=border, bottom=border)

        title_style.border = borders
        default_style.border = borders

        def create_title(ws, titles):
            if len(titles) > 3:
                ws.column_dimensions[get_column_letter(2)].width = 30
            for i, title in enumerate(titles, 1):
                ws.cell(1,i).value = title
                ws.cell(1,i).style = title_style

        def fill_row(ws, row_id, row_data, full=False):
            for i, cell in enumerate(row_data[:5], 1):
                if not full:
                    if i == 2:
                        continue
                    elif i > 2:
                        i = i-1
                ws.cell(row_id, i).value = cell
                ws.cell(row_id, i).style = default_style


        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Наличие"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Наличие"
        wb3 = Workbook()
        ws3 = wb3.active
        ws3.title = "Наличие"

        create_title(ws1, ("Артикул", "Партия", "Доступно"))
        create_title(ws2, ("Артикул", "Партия", "Доступно"))
        create_title(ws3, ("Артикул", "Наименование", "Партия", "Доступно"))

        alt_i = 2
        for i, row in enumerate(data_list, 2):
            fill_row(ws2, i, row)
            if row[3]:
                fill_row(ws1, alt_i, row)
                fill_row(ws3, alt_i, row, full=True)
                alt_i += 1

        wb1.save(dir_path + "%s.xlsx" % filename)
        wb2.save(dir_path + "%s с отсуствующими.xlsx" % filename)
        wb3.save(dir_path + "%s с наименованиями.xlsx" % filename)


    # codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    codes_name = names_session.query(VCodeName).all()
    # _collections = names_session.query(Collection_Factory).all()
    factories = names_session.query(Factory).all()

    _factories_collection_pair = {'excl': []}
    
    for factory in factories:
        _factories_collection_pair[factory.name] = []
        for coll in collections.copy():
            for f_coll in factory.collections:
                if coll.name == f_coll.name:
                    _excc = False
                    for _exception in exclusive_collections:
                        if not coll.name:
                            break
                        if _exception.lower() in coll.name.lower():
                            _excc = True
                            break
                    if _excc:
                        _factories_collection_pair['excl'].append(coll)
                    else:
                        _factories_collection_pair[factory.name].append(coll)
                    collections.remove(coll)

    code_names_dict = {}
    for code in codes_name:
        code_names_dict[code.code] = code.name

    row_list = get_vcodes_list("DID")
    create_xlsx_table("DID", row_list)

    row_list = get_vcodes_list("G'BOYA")
    create_xlsx_table("G'BOYA", row_list)

    row_list_ = get_vcodes_list("YIEN")
    create_xlsx_table("YIEN", row_list_)

    row_list += row_list_
    create_xlsx_table("Китай", row_list)

    row_list = get_vcodes_list("Wiganford")
    create_xlsx_table("Wiganford", row_list)

    # row_list = get_vcodes_list("Rasch")
    # create_xlsx_table("Rasch", row_list)

    row_list = get_vcodes_list("Partner")
    create_xlsx_table("Фотообои", row_list)

    row_list = get_vcodes_list("Emiliana Parati")
    row_list += get_vcodes_list("Limonta")
    row_list += get_vcodes_list("Parato")
    create_xlsx_table("Италия", row_list)

    row_list = get_vcodes_list("Индустрия")
    row_list += get_vcodes_list("Элизиум")
    create_xlsx_table("Россия", row_list)

    row_list = get_vcodes_list("excl")
    create_xlsx_table("RC VY Ferre", row_list)

def create_xlsx_without_consigments(dir_path, session):
    filename = 'Наличие без партий'
    
    title_style = NamedStyle(name="title")
    title_style.fill = PatternFill("solid", fgColor="00F4ECC5")
    title_style.font = Font(name='Arial', sz=10.0)

    default_style = NamedStyle(name="default")
    default_style.font = Font(name='Arial', sz=8.0)

    border = Side(style='thin', color='00CCC085')
    borders = Border(left=border, top=border, right=border, bottom=border)

    title_style.border = borders
    default_style.border = borders

    def create_title(ws, titles):

        ws.cell(2,1).value = "Остатки и доступность товара"
        ws.cell(2,1).font = Font(name='Arial', sz=18.0, bold=True)
        ws.cell(5,1).value = "Склад"
        ws.cell(5,1).style = title_style
        ws.cell(5,2).value = "Сейчас"
        ws.cell(5,2).style = title_style
        if len(titles) > 3:
            ws.column_dimensions[get_column_letter(2)].width = 30
        for i, title in enumerate(titles, 1):
            ws.cell(6,i).value = title
            ws.cell(6,i).style = title_style


    wb = Workbook()
    ws = wb.active
    ws.title = "Наличие"
    create_title(ws, ("Артикул", "Доступно"))

    vcodes = session.query(VCode).all()
    data_list = []
    for vcode in vcodes:
        if vcode.consigments:
            amount = 0
            for consig in vcode.consigments:
                amount += consig.amount
            data_list.append((vcode.code, amount))
        # else:
        #     data_list.append((vcode.code, 0))

    for r, row_data in enumerate(data_list, 7):  # 7 - table start row
        for c, cell in enumerate(row_data, 1):
            ws.cell(r, c).value = cell
            ws.cell(r, c).style = default_style


    wb.save(dir_path + "%s.xlsx" % filename)
    return True


def send_mails_exchange(dir_path, lg, ps, sv, mac, mail_list, mary=False):
    # dsd = base64.b64decode(dsd).decode()
    # sss = base64.b64decode(sss).decode()
    credentials = Credentials(lg, ps)

    # lsk = base64.b64decode(lsk).decode()
    # msk = base64.b64decode(msk).decode()
    config = Configuration(server=sv, credentials=credentials)
    account = Account(primary_smtp_address=mac, config=config,
                      autodiscover=False, access_type=DELEGATE)
    m1 = Message(
        account=account,
        subject='Наличие',
        body='',
        to_recipients=mail_list,
    )
    m2 = Message(
        account=account,
        subject='Наличие с наименованиями',
        body='',
        to_recipients=[mail_list[0]],
    )
    m3 = Message(
        account=account,
        subject='Наличие со всеми артикулами',
        body='',
        to_recipients=[mail_list[0]],
    )

    for filename in os.listdir(dir_path):
        if 'отсуствующими' in filename:
            m3.attach(FileAttachment(
                name=filename,
                content=open(dir_path + filename, 'rb').read(),
            ))
        elif 'наименованиями' in filename:
            m2.attach(FileAttachment(
                name=filename,
                content=open(dir_path + filename, 'rb').read(),
            ))
        else:
            m1.attach(FileAttachment(
                name=filename,
                content=open(dir_path + filename, 'rb').read(),
            ))
    if not mary:
        m1.attach(FileAttachment(
            name='Наличие.xlsx',
            content=open(xlxs_filepath, 'rb').read(),
        ))
    # m1.attach(FileAttachment(
    #     name='Наличие без партий.xlsx',
    #     content=open(dir_path + 'Наличие без партий.xlsx', 'rb').read(),
    # ))

    m1.send()
    if not mary:
        m2.send()
        m3.send()


def create_amounts_to_mary(dir_path, session):

    filename = 'Наличие'
    
    title_style = NamedStyle(name="title")
    title_style.fill = PatternFill("solid", fgColor="00F4ECC5")
    title_style.font = Font(name='Arial', sz=10.0)

    default_style = NamedStyle(name="default")
    default_style.font = Font(name='Arial', sz=8.0)

    border = Side(style='thin', color='00CCC085')
    borders = Border(left=border, top=border, right=border, bottom=border)

    title_style.border = borders
    default_style.border = borders

    def create_title(ws, titles):

        ws.cell(2,1).value = "Остатки и доступность товара"
        ws.cell(2,1).font = Font(name='Arial', sz=18.0, bold=True)
        ws.cell(5,1).value = "Склад"
        ws.cell(5,1).style = title_style
        ws.cell(5,2).value = "Характеристика"
        ws.cell(5,2).style = title_style
        ws.cell(5,3).value = "Сейчас"
        ws.cell(5,3).style = title_style
        if len(titles) > 3:
            ws.column_dimensions[get_column_letter(2)].width = 30
        for i, title in enumerate(titles, 1):
            ws.cell(6,i).value = title
            ws.cell(6,i).style = title_style


    wb = Workbook()
    ws = wb.active
    ws.title = "Наличие"
    create_title(ws, ("Артикул", " ", "Доступно"))

    vcodes = session.query(VCode).all()
    data_list = []
    for vcode in vcodes:
        if vcode.consigments:
            for consig in vcode.consigments:
                data_list.append((vcode.code, consig.name, consig.amount))
        # else:
        #     data_list.append((vcode.code, 0))

    for r, row_data in enumerate(data_list, 7):  # 7 - table start row
        for c, cell in enumerate(row_data, 1):
            ws.cell(r, c).value = cell
            ws.cell(r, c).style = default_style


    wb.save(dir_path + "%s.xlsx" % filename)
    return True


def send_mails_smtp(dir_path, send_from, server, mail_list, mary=False, text=''):

    def send_mail(send_from, send_to, subject, text, files=None, server="127.0.0.1"):
        assert isinstance(send_to, list)

        msg = MIMEMultipart()
        msg['From'] = send_from
        msg['To'] = COMMASPACE.join(send_to)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = subject

        msg.attach(MIMEText(text))

        if type(files) == dict:
            for fn in files:
                with open(files[fn], "rb") as file:
                    part = MIMEApplication(
                        file.read(),
                        Name=fn
                    )
                # After the file is closed
                part['Content-Disposition'] = 'attachment; filename="%s"' % fn
                msg.attach(part)
        else:
            for f in files or []:
                with open(f, "rb") as file:
                    part = MIMEApplication(
                        file.read(),
                        Name=os.path.basename(f)
                    )
                # After the file is closed
                part['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(f)
                msg.attach(part)


        smtp = smtplib.SMTP(server)
        smtp.sendmail(send_from, send_to, msg.as_string())
        smtp.close()


    filelists = [{}, {}, {}]
    for filename in os.listdir(dir_path):
        if 'отсуствующими' in filename:
            filelists[0][filename] = os.path.join(dir_path, filename)
        elif 'наименованиями' in filename:
            filelists[1][filename] = os.path.join(dir_path, filename)
        else:
            filelists[2][filename] = os.path.join(dir_path, filename)
    if not mary:
        filelists[0]['Наличие.xlsx'] = xlxs_filepath

    send_mail(send_from, mail_list, 'Наличие', text, filelists[0], server)
    if not mary:
        send_mail(send_from, mail_list, 'Наличие с наименованиями', text, filelists[1], server)
        send_mail(send_from, mail_list, 'Наличие со всеми артикулами', text, filelists[2], server)
