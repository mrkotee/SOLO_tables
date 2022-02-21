from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError, InvalidRequestError
from alchemy_models import VCode, Consigment, Collection, base_path, Table_row
from openpyxl import load_workbook
import random, string, time
import os


def read_xlxs_uni(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[5].value.split()[0]
        motive = row[2].value
        if motive == "мотив":
            motive = True
        else:
            motive = False
        xlxs_rows.append((i, vcode, motive))
    wb.close()
    return xlxs_rows


def migrate_vcodes():
    # xlsx_rows_uni = read_xlxs_uni(filepath)

    from models_old import VCode as VCode_old
    codes_old = session.query(VCode_old).all()

    for vcode_old in codes_old:
        print(vcode_old.id)
        # print(vcode_old.__dict__)

        new_vcode = VCode(code=vcode_old.code, collection_id=vcode_old.collection_id, motive=vcode_old.motive)
        # for row in xlsx_rows_uni:
        #     if row[1] == vcode_old.code:
        #         new_vcode.motive = row[2]
        session_t.add(new_vcode)

        session_t.commit()

        print(new_vcode.id)
        print(new_vcode.__dict__)


def migrate_consig():
    from models_old import Consigment as Consigment_old
    consigs_old = session.query(Consigment_old).all()

    for consig_old in consigs_old:
        print(consig_old.id)
        new_consig = Consigment(consig_old.name, consig_old.vcode_id, consig_old.amount)

        session_t.add(new_consig)

        session_t.commit()
        print(new_consig.id)


def migrate_collection():
    from models_old import Collection as Collection_old
    collections_old = session.query(Collection_old).all()

    for collection_old in collections_old:
        print(collection_old.id)
        new_collection = Collection(collection_old.name, collection_old.boxes)

        session_t.add(new_collection)

        session_t.commit()
        print(new_collection.id)


if __name__ == '__main__':
    # base_filename = "base_test.bd"
    # base_path = os.path.join(os.getcwd(), base_filename)
    # t_base_path = os.path.join(os.getcwd(), base_filename + '_temp')
    t_base_path = base_path + '_temp'
    # base_path = r'/home/django/bike_shop/solo/base_test.bd'
    # filepath = r'/home/django/bike_shop/solo/base.xlsx'
    # filepath = os.path.join(os.getcwd(), "base.xlsx")
    filepath = os.path.join(os.getcwd(), "uni-motiv.xlsx")

    if not os.path.exists(t_base_path):
        from alchemy_models import Base
        engine_t = create_engine('sqlite:///%s' % t_base_path, echo=False)
        Base.metadata.create_all(bind=engine_t)

    engine_t = create_engine('sqlite:///%s' % t_base_path, echo=False)
    Session_t = sessionmaker(bind=engine_t)
    session_t = Session_t()

    engine = create_engine('sqlite:///%s' % base_path, echo=False)
    Session = sessionmaker(bind=engine)
    session = Session()


    migrate_collection()

    migrate_vcodes()

    migrate_consig()



    session.close()
    session_t.close()

    os.rename(base_path, base_path+'_old')
    os.rename(t_base_path, base_path)
    print(base_path+"_old", "created")
