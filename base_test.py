
import random, string, time
import os
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import QueuePool
from sqlalchemy.exc import IntegrityError, InvalidRequestError
from models import VCode, Consigment, Collection, base_path, Table_row
from solo_settings import replace_dict
from openpyxl import load_workbook



def randomword(length):
   letters = string.ascii_lowercase
   return ''.join(random.choice(letters) for i in range(length))


def add_to_base():
    codes = ["E37108", 'E37100', 'E37105']

    for code in codes:
        code_in_base = session.query(VCode).filter(VCode.code == code).all()
        # print(code_in_base)
        if not code_in_base:
            session.add(VCode(code))

    for code in codes:
        code_in_base = session.query(VCode).filter(VCode.code == code).first()

        if random.randint(1, 3) == 2:
            name = random.randint(10358, 20067)
        else:
            name = randomword(4)

        consig = Consigment(name=name, vcode=code_in_base.id, amount=random.randint(6, 36))
        consig_in_base = session.query(Consigment).filter(
                                                    Consigment.vcode_id == code).filter(
                                                        Consigment.name == name).all()
        if not consig_in_base:
            session.add(consig)

    session.commit()

    for code in codes:
        code_in_base = session.query(VCode).filter(VCode.code == code).first()
        for consig in code_in_base.consigments:
            print(code_in_base.code, consig.name, consig.amount)


def read_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[0].value.split()[0]
        consig = row[1].value
        amount = row[2].value
        xlxs_rows.append((i, vcode, consig, amount))
    wb.close()
    return xlxs_rows


def read_abc_xlxs(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    xlxs_rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i < 6:
            continue
        vcode = row[0].value
        name = row[1].value
        collection = row[2].value
        box = row[3].value
        xlxs_rows.append((vcode, name, collection, box))
    wb.close()
    return xlxs_rows


def get_all_from_base(xlxs_rows):
    codes = session.query(VCode).all()
    consigs = session.query(Consigment).all()

    need_commit = False
    for row in xlxs_rows:
        
        i = row[0]
        vcode = row[1]
        consig = row[2]
        amount = row[3]

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                # codes.remove(_code)
                break
        if not code_in_base:
            session.add(VCode(vcode))
            session.commit()
            # need_commit = True
            code_in_base = session.query(VCode).filter(VCode.code == vcode).first()
            codes.append(code_in_base)

        consig_in_base = None
        for _consig in code_in_base.consigments:
            if _consig.name == consig:
                consig_in_base = _consig
                consigs.remove(_consig)
                break 
        if not consig_in_base:
            _consig = Consigment(name=consig, vcode=code_in_base.id, amount=amount)
            session.add(_consig)
            # session.commit()
            need_commit = True
            consig_in_base = session.query(Consigment).filter(
                                Consigment.vcode_id == code_in_base.id).filter(
                                    Consigment.name == consig).first()

        if amount != consig_in_base.amount:
            print('change {} amount from {} to {}'.format(consig_in_base.name, consig_in_base.amount, amount))
            consig_in_base.amount = amount
            need_commit = True

        if need_commit and not i % 100:
            session.commit()
            need_commit = False
        # if code_in_base.collection:
        #     print(i, code_in_base.collection.name, code_in_base.code, consig_in_base.name, consig_in_base.amount)
        #     print(consig_in_base.amount, amount)
        # else:
        #     print(i, "no collection", code_in_base.code, consig_in_base.name)


    print("rows_to_check: ", len(consigs))
    print("checked rows: ", i)
    for consig_in_base in consigs:
        # print('delete this consigments:')
        print('zero amount to this consigments')
        print(session.query(VCode).get(consig_in_base.vcode_id).code, consig_in_base.name)
        consig_in_base.amount = 0
        # session.delete(consig_in_base)
    session.commit()


def get_one_by_one():
    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[0].value
        consig = row[1].value
        amount = row[2].value

        code_in_base = session.query(VCode).filter(VCode.code == vcode).first()
        if not code_in_base:
            session.add(VCode(vcode))
            session.commit()
            code_in_base = session.query(VCode).filter(VCode.code == vcode).first()

        consig_in_base = session.query(Consigment).filter(
                            Consigment.vcode_id == code_in_base.id).filter(
                                Consigment.name == consig).first()
        if not consig_in_base:
            _consig = Consigment(name=consig, vcode=code_in_base.id, amount=amount)
            session.add(_consig)
            session.commit()
            consig_in_base = session.query(Consigment).filter(
                                Consigment.vcode_id == code_in_base.id).filter(
                                    Consigment.name == consig).first()

        if amount != consig_in_base.amount:
            consig_in_base.amount = amount
    wb.close()
    session.commit()


def add_boxes_to_vcodes(xlxs_rows, session):
    """Выгрузка из ABC"""

    def find_collection(collection_name, collections):
        for _coll in collections:
            if _coll.name == collection_name:
                return _coll
        return None

    codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = str(row[0])
        name = row[1]
        collection = row[2]
        box = row[3]

        collection_in_base = find_collection(collection, collections)

        if collection_in_base and collection_in_base.id != 125:
            if box and box != collection_in_base.boxes:
                collection = collection + "_" + str(box)
                collection_in_base = find_collection(collection, collections)

        if not collection_in_base:
            response.append("added collection %s" % collection)
            session.add(Collection(collection, box))
            session.commit()
            collection_in_base = session.query(Collection).filter(Collection.name == collection).first()
            collections.append(collection_in_base)

        code_in_base = None
        for _code in codes:
            # if _code.code == '81740':
            #     print(len(str(vcode)))
            #     print(len(_code.code))
            #     print(_code.code == vcode)
            if _code.code == vcode:
                code_in_base = _code
                if not code_in_base.collection_id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("added collection to %s" % vcode)
                elif int(code_in_base.collection_id) != collection_in_base.id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("changed collection to %s" % vcode)
                break
        if not code_in_base:
            response.append("added vcode %s" % vcode)
            session.add(VCode(vcode, collection_in_base.id))
            need_commit = True

        if not i % 100 and need_commit:
            session.commit()
            need_commit = False

    if need_commit:
        session.commit()

    if not response:
        response.append("no changes")
        
    return response


def add_vcodes():
    codes = session.query(VCode).all()

    wb = load_workbook(filename=filepath)
    ws = wb[wb.sheetnames[0]]
    need_commit = False
    prev_vcode = ''
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        vcode = row[0].value
        # consig = row[1].value
        # amount = row[2].value
        if vcode == prev_vcode:
            continue
        else:
            prev_vcode = vcode
        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                # codes.remove(_code)
                break
        if not code_in_base:
            session.add(VCode(vcode))
            # session.commit()
            need_commit = True
            print(i, 'added %s' % vcode)
            # code_in_base = session.query(VCode).filter(VCode.code == vcode).first()
            # codes.append(code_in_base)

        if need_commit and not i % 100:
            session.commit()
            need_commit = False

    wb.close()
    session.commit()


def get_for_table(data_str, all_boxes_num=0, uni_boxes_num=0):
    data_list = data_str.split()

    for_range = range(len(data_list))

    table_rows = []
    skip = False
    in_boxes = False
    boxes_num = 0
    for i in for_range:
        if skip:
            skip = False
            continue

        vcode = data_list[i]

        if len(vcode) < 4 and not vcode[0].isdigit():
            data_list[i+1] = vcode + data_list[i+1]
            continue

        if i == for_range.stop-1 or len(data_list[i+1]) > 3:
            in_boxes = True
            # skip = True
        elif "*" in data_list[i+1]:
            in_boxes = True
            skip = True
            # boxes_num = int(data_list[i+1][1:])
            boxes_num = data_list[i+1].replace("*", "")
        elif len(data_list[i+1]) < 4 and not data_list[i+1][0].isdigit():
            skip = True
            data_list[i+2] = data_list[i+1] + data_list[i+2]
        else:
            in_boxes = False
            skip = True

        if not in_boxes:
            number = data_list[i + 1]
        else:
            number = 0

        # if len(vcode) < 4 and len(str(number)) > 3:
        #     vcode, number = number, vcode

        number = int(number)
        vcode = vcode.upper()
        for lett in vcode:
            if lett in replace_dict:
                vcode = vcode.replace(lett, replace_dict[lett])
        row = Table_row(vcode, number)
        table_rows.append(row)

    # for row in table_rows:
        # print('trying %s\n' % row.vcode)
        code_in_base = session.query(VCode).filter(VCode.code == row.vcode).first()
        _code_in_base = None
        if not code_in_base:
            check_for_photo(session, vcode, row)
                    # continue
            code_in_base = find_code(vcode, session)
            if not code_in_base:
                row.consig = ""
                row.comment += "Артикул не найден "
                continue

            elif len(code_in_base) > 1:
                row.comment += "арт. %s не найден " % row.vcode
                minimal = min([code.code for code in code_in_base], key=len) # самый короткий артикул
                row.comment += "Возможные артикулы: "
                for _code in code_in_base:
                    row.comment += "{} ".format(_code.code)
                    if _code.code == minimal:
                        _code_in_base = _code   # выбор самого короткого артикула
                    # row.vcode = code_in_base[0].code
                    continue
            else:
                # row.comment += "Возможно нужен %s " % code_in_base[0].code
                row.vcode = code_in_base[0].code
                code_in_base = code_in_base[0]
            # continue

        if _code_in_base:
            code_in_base = _code_in_base
            row.vcode = code_in_base.code

        if in_boxes:
            if code_in_base.collection and not check_for_photo(session, vcode, row):
                if boxes_num:
                    row.number = code_in_base.collection.boxes * int(boxes_num)
                    boxes_num = 0
                elif uni_boxes_num and not code_in_base.motive:
                    row.number = code_in_base.collection.boxes * int(uni_boxes_num)
                elif all_boxes_num:
                    row.number = code_in_base.collection.boxes * int(all_boxes_num)
                else:
                    row.number = code_in_base.collection.boxes
            else:
                row.number = 1
        if code_in_base.consigments:
            enough = False
            for consig in code_in_base.consigments:
                if consig.amount >= row.number:
                    enough = True
                    if consig.name:
                        row.consig = consig.name
                    break
            if not enough:
                row.comment += "Недостаточно на св. остатке."
                add_comment = " в наличии есть партии: "
                max_num_consig = None
                max_num = 0
                for consig in code_in_base.consigments:
                    if consig.amount >= 0:
                        if consig.amount > max_num:
                            max_num_consig = consig.name
                            max_num = consig.amount
                        add_comment += "{} {} рул. ".format(consig.name, consig.amount)
                row.consig = max_num_consig
                if len(add_comment) > 24:
                    row.comment += add_comment
        else:
            row.comment = "Партий не найдено"
            if not check_for_photo(session, vcode, row):
                row.consig = "Общая"
                
            else:
                continue

        check_for_photo(session, vcode, row)

    return table_rows


def find_code(vcode, session):
    code_in_base = session.query(VCode).filter(VCode.code.contains(vcode)).all()

    if not code_in_base:
        return False
    len_vcode = len(vcode)
    for i, code in enumerate(code_in_base.copy()):
        if code.code[0].isdigit():  # len(code.code) > len_vcode+2 and 
            code_in_base.remove(code)
    # if len(code_in_base) > 1:
    #     return min([code.code for code in code_in_base], key=len)
    return code_in_base


def find_analog_photowp(vcode, session):
    digit_code = ""
    for lett in vcode:
        if lett.isdigit():
            digit_code += lett
        else:
            break
    response = find_code(digit_code, session)
    if not response:
        return False
    codes = []
    for _code in response.copy():
        if "V" in _code.code or "P" in _code.code:
            if len(_code.consigments) > 0 and _code.consigments[0].amount > 0:
                codes.append(_code)
    return codes

def check_for_photo(session, vcode, row):
    if "V" in vcode or "P" in vcode:
        check_photo = find_analog_photowp(vcode, session)
        if check_photo and len(check_photo) > 1:
            row.comment += "Аналоги: "
            for code in check_photo:
                if code.code == vcode:
                    continue
                row.comment += code.code + " "
                row.comment += str(code.consigments[0].amount) + " шт, "
            row.comment = row.comment[:-2]
            if row.number == 0:
                row.number = 1
        return True
    return False


def find_duplicates():
    """find and delete duplicated vcodes in base"""
    codes = session.query(VCode).all()
    deleted_id = []
    # print(len(codes), len(codes.copy()))
    for code in codes.copy():
        # print("checking %s" % code.code)
        code_in_base = session.query(VCode).filter(VCode.code.contains(code.code)).all()

        if len(code_in_base) > 1:
            # print(code_in_base)
            for c in code_in_base:
                if c.id == code.id:
                    continue
                if c.code.split()[0] == code.code.split()[0]:
                    print(c.id, c.code)
                    print(code.id, code.code)
                    # print("id %d code %s" % (c.id, c.code))
                    print("consigments %s" % c.consigments)
                    if c.id > code.id:
                        for consig in c.consigments:
                            consig.vcode_id = code.id
                            session.commit()
                        session.delete(c)
                        session.commit()
                        print("deleted")


def check_collection(xlxs_rows, session):
    codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = str(row[0])
        name = row[1]
        collection = row[2]
        box = row[3]

        collection_in_base = None
        for _coll in collections:
            if _coll.name == collection:
                collection_in_base = _coll
                break

        if collection_in_base.id == 125:
            continue

        if box and box != collection_in_base.boxes:
            print(vcode, collection, box, collection_in_base.boxes)


def change_collection_boxes(xlxs_rows, session):

    def find_collection(collection_name, collections):
        for _coll in collections:
            if _coll.name == collection_name:
                return _coll
        return None

    codes = session.query(VCode).all()
    collections = session.query(Collection).all()

    response = []
    need_commit = False
    for i, row in enumerate(xlxs_rows):
        vcode = str(row[0])
        name = row[1]
        collection = row[2]
        box = row[3]

        collection_in_base = None
        collection_in_base = find_collection(collection, collections)

        if collection_in_base.id == 125:
            continue

        if box and box != collection_in_base.boxes:
            new_collection_name = collection + "_" + str(box)
            collection_in_base = find_collection(new_collection_name, collections)
            if not collection_in_base:
                response.append("added collection %s" % new_collection_name)
                session.add(Collection(new_collection_name, box))
                session.commit()
                collection_in_base = session.query(Collection).filter(Collection.name == new_collection_name).first()
                collections.append(collection_in_base)

        code_in_base = None
        for _code in codes:
            if _code.code == vcode:
                code_in_base = _code
                if not code_in_base.collection_id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("added collection to %s" % vcode)
                elif int(code_in_base.collection_id) != collection_in_base.id:
                    code_in_base.collection_id = collection_in_base.id
                    need_commit = True
                    response.append("changed collection to %s" % vcode)
                break

    if need_commit:
        session.commit()

    return(response)


if __name__ == '__main__':
    base_filename = "base_test.bd"
    base_path = os.path.join(os.getcwd(), base_filename)
    # base_path = r'/home/django/bike_shop/solo/base_test.bd'
    # filepath = r'/home/django/bike_shop/solo/base.xlsx'
    filepath = os.path.join(os.getcwd(), "base.xlsx")

    engine = create_engine('sqlite:///%s' % base_path, echo=False, poolclass=QueuePool)
    Session = sessionmaker(bind=engine)
    session = Session()



    ######  speed test ########
    # start = time.time()
    # get_one_by_one()
    # end = time.time()
    # print("get_one_by_one done in {} sec".format(end-start))

    # start = time.time()
    # xlxs_rows = read_xlxs(filepath)
    # get_all_from_base(xlxs_rows)
    # end = time.time()
    # print("get_all_from_base done in {} sec".format(end-start))

    ############################


    # add_to_base()


    # filepath = r'/home/django/bike_shop/solo/base.xlsx'
    # add_vcodes()



    #######  abc tests ########

    # xlxs_abc_filepath = r'/home/django/bike_shop/solo/abc.xlsx'
    # xlxs_abc_filepath = r'/home/django/bike_shop/solo/abc_test.xlsx'
    # xlxs_rows = read_abc_xlxs(xlxs_abc_filepath)
    # changed_positions = add_boxes_to_vcodes(xlxs_rows, session)


    # check_collection(xlxs_rows, session)
    # changed_positions = change_collection_boxes(xlxs_rows, session)

    # for change in changed_positions:
    #     print(change)

    ###########################




    ###### table tests ########

    request_str = "rmg 2301-1 EL21201 2 e37105 *2 N55664 4 167062-90 894P8 136P8  RMG2303-1  2303-1  ak 20115 ак2031 2050 ak2050"
    # request_str = "21201 37108 *20 55664 167062-90 894p8 136p4"
    # request_str = "240509 240461"
    response = get_for_table(request_str, all_boxes_num=3, uni_boxes_num=6)
    for row in response:
        print(row.vcode, row.consig, row.number, row.comment)


    #############################



    # find_duplicates()





    session.close()
